"""Lee los valores reales de las columnas clave de la primera fila de datos."""
import openpyxl
import pandas as pd

RUTA = r"D:\Proyecto2026\ACP_DWH\ACP Proyecciones\39. Proyecciones Semana 5 + 5 semanas (Conteo S3).xlsx"

wb = openpyxl.load_workbook(RUTA, data_only=True)
ws = wb["Calculo"]

# Primera fila de datos reales es fila 5
# Columnas clave identificadas:
# Col 2: Sem (semana ISO de evaluacion)
# Col 4: M (Modulo)
# Col 5: T (Turno)
# Col 6: V (Valvula)
# Col 7: Variedad
# Col 8: Area
# Col 9: Plantas (total)
# Col 15-23: Conteo por estado (Boton, Flor, Pequena, Verde, F1, F2, Cremas, Maduras, Cosechable)
# Col 24-63: % Maduracion por estado x semana (la MATRIZ)
# Col 64-74: Peso Baya kg Sem 1 a Sem 11
# Col 75-85: % Plantas Productivas Sem 1 a Sem 11
# Col 86-96: Kg calculados Sem 1 a Sem 11

print("=== VALORES REALES PRIMERAS 5 FILAS DE DATOS (fila 5 en adelante) ===\n")

cols_interes = {
    2: "Semana_ISO",
    4: "Modulo",
    5: "Turno",
    6: "Valvula",
    7: "Variedad",
    9: "Plantas_Total",
    15: "Cont_BotonFloral",
    16: "Cont_Flor",
    17: "Cont_Pequena",
    18: "Cont_Verde",
    19: "Cont_Fase1",
    20: "Cont_Fase2",
    21: "Cont_Cremas",
    22: "Cont_Maduras",
    23: "Cont_Cosechable",
    # % Maduracion Sem 1
    24: "PctMad_F1_S1",
    25: "PctMad_F2_S1",
    26: "PctMad_Cremas_S1",
    27: "PctMad_Maduras_S1",
    28: "PctMad_Cosechable_S1",
    # % Maduracion Sem 2
    29: "PctMad_F1_S2",
    30: "PctMad_F2_S2",
    31: "PctMad_Cremas_S2",
    32: "PctMad_Maduras_S2",
    33: "PctMad_Cosechable_S2",
    # % Maduracion Sem 3
    34: "PctMad_Verde_S3",
    35: "PctMad_F1_S3",
    36: "PctMad_F2_S3",
    # Peso Baya por semana
    64: "PesoBaya_S1",
    65: "PesoBaya_S2",
    66: "PesoBaya_S3",
    67: "PesoBaya_S4",
    68: "PesoBaya_S5",
    69: "PesoBaya_S6",
    # % Productivas
    75: "PctProd_S1",
    76: "PctProd_S2",
    77: "PctProd_S3",
    78: "PctProd_S4",
    79: "PctProd_S5",
    80: "PctProd_S6",
    # Kg proyectados
    86: "Kg_S1",
    87: "Kg_S2",
    88: "Kg_S3",
    89: "Kg_S4",
    90: "Kg_S5",
    91: "Kg_S6",
}

records = []
for fila in range(5, 20):
    rec = {"Fila": fila}
    tiene_dato = False
    for col, nombre in cols_interes.items():
        val = ws.cell(fila, col).value
        rec[nombre] = val
        if val is not None:
            tiene_dato = True
    if tiene_dato:
        records.append(rec)
    if len(records) >= 5:
        break

df = pd.DataFrame(records)
print(df.T.to_string())

print("\n\n=== ANÁLISIS DE LÓGICA DE PESOS ===")
print("En el Excel, ¿el Peso Baya S1 es de la semana de EVALUACION o de la SIGUIENTE?")
print()
for fila in range(5, 12):
    sem = ws.cell(fila, 2).value
    mod = ws.cell(fila, 4).value
    var = ws.cell(fila, 7).value
    peso_s1 = ws.cell(fila, 64).value
    peso_s2 = ws.cell(fila, 65).value
    if sem and mod and var:
        print(f"  Fila {fila}: Sem_Eval={sem}, Modulo={mod}, Variedad={var}")
        print(f"    PesoBaya_S1={peso_s1}, PesoBaya_S2={peso_s2}")
        print()
