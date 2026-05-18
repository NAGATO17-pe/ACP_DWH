"""Análisis detallado del Excel de proyecciones - hoja Calculo y Pesos."""
import openpyxl
import pandas as pd

RUTA = r"D:\Proyecto2026\ACP_DWH\ACP Proyecciones\39. Proyecciones Semana 5 + 5 semanas (Conteo S3).xlsx"

wb = openpyxl.load_workbook(RUTA, data_only=True)
print(f"Hojas disponibles: {wb.sheetnames}\n")

# ── Hoja Calculo ─────────────────────────────────────────────────────────────
print("=" * 80)
print("HOJA: Calculo (primeras 80 filas, primeras 30 columnas)")
print("=" * 80)
ws = wb["Calculo"]
rows = []
for row in ws.iter_rows(min_row=1, max_row=80, max_col=30, values_only=True):
    rows.append(row)
df = pd.DataFrame(rows)
# Eliminar columnas completamente vacías
df_clean = df.dropna(how='all', axis=1)
print(df_clean.to_string())
print()

# ── Hoja Pesos ───────────────────────────────────────────────────────────────
if "Pesos" in wb.sheetnames:
    print("=" * 80)
    print("HOJA: Pesos (primeras 50 filas)")
    print("=" * 80)
    ws2 = wb["Pesos"]
    rows2 = []
    for row in ws2.iter_rows(min_row=1, max_row=50, max_col=20, values_only=True):
        rows2.append(row)
    df2 = pd.DataFrame(rows2).dropna(how='all', axis=1).dropna(how='all', axis=0)
    print(df2.to_string())
    print()

# ── Hoja Conteo ──────────────────────────────────────────────────────────────
for sheet_name in wb.sheetnames:
    if "onteo" in sheet_name or "onteos" in sheet_name:
        print("=" * 80)
        print(f"HOJA: {sheet_name} (primeras 30 filas)")
        print("=" * 80)
        ws3 = wb[sheet_name]
        rows3 = []
        for row in ws3.iter_rows(min_row=1, max_row=30, max_col=25, values_only=True):
            rows3.append(row)
        df3 = pd.DataFrame(rows3).dropna(how='all', axis=1).dropna(how='all', axis=0)
        print(df3.to_string())
        print()
        break

# ── Hoja Plantillas / Productivas ────────────────────────────────────────────
for sheet_name in wb.sheetnames:
    if "lanta" in sheet_name or "elada" in sheet_name or "Censo" in sheet_name:
        print("=" * 80)
        print(f"HOJA: {sheet_name} (primeras 30 filas)")
        print("=" * 80)
        ws4 = wb[sheet_name]
        rows4 = []
        for row in ws4.iter_rows(min_row=1, max_row=30, max_col=20, values_only=True):
            rows4.append(row)
        df4 = pd.DataFrame(rows4).dropna(how='all', axis=1).dropna(how='all', axis=0)
        print(df4.to_string())
        print()
        break
