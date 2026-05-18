"""Buscar filas del Excel donde haya valores NO CERO en Pesos y Kg para entender la lógica."""
import openpyxl
import pandas as pd

RUTA = r"D:\Proyecto2026\ACP_DWH\ACP Proyecciones\39. Proyecciones Semana 5 + 5 semanas (Conteo S3).xlsx"

wb = openpyxl.load_workbook(RUTA, data_only=True)
ws = wb["Calculo"]

print(f"Total filas: {ws.max_row}, Total cols: {ws.max_column}\n")

# Buscar fila donde Kg_S1 (col 86) sea distinto de 0 y no None
print("=== FILAS CON Kg_S1 > 0 ===")
count = 0
for fila in range(5, ws.max_row + 1):
    kg_s1 = ws.cell(fila, 86).value
    if kg_s1 and kg_s1 != 0:
        sem = ws.cell(fila, 2).value
        mod = ws.cell(fila, 4).value
        turno = ws.cell(fila, 5).value
        valv = ws.cell(fila, 6).value
        var = ws.cell(fila, 7).value
        plantas = ws.cell(fila, 9).value
        # Conteos
        cont_f1 = ws.cell(fila, 19).value
        cont_f2 = ws.cell(fila, 20).value
        cont_cremas = ws.cell(fila, 21).value
        cont_maduras = ws.cell(fila, 22).value
        cont_cosech = ws.cell(fila, 23).value
        # Pesos
        peso_s1 = ws.cell(fila, 64).value
        peso_s2 = ws.cell(fila, 65).value
        peso_s3 = ws.cell(fila, 66).value
        # % Productivas
        prod_s1 = ws.cell(fila, 75).value
        # % Maduracion Sem 1
        pct_cos_s1 = ws.cell(fila, 28).value
        pct_mad_s1 = ws.cell(fila, 27).value
        pct_crem_s1 = ws.cell(fila, 26).value
        # Kg
        kg_s2 = ws.cell(fila, 87).value
        kg_s3 = ws.cell(fila, 88).value

        print(f"\nFila {fila}: Sem={sem}, M={mod}, T={turno}, V={valv}, Var={var}")
        print(f"  Plantas={plantas}")
        print(f"  Conteo: F1={cont_f1}, F2={cont_f2}, Cremas={cont_cremas}, Maduras={cont_maduras}, Cosechable={cont_cosech}")
        print(f"  % Mad S1: Cremas={pct_crem_s1}, Maduras={pct_mad_s1}, Cosechable={pct_cos_s1}")
        print(f"  PesoBaya (kg): S1={peso_s1}, S2={peso_s2}, S3={peso_s3}")
        print(f"  %Productivas S1={prod_s1}")
        print(f"  => Kg: S1={kg_s1:.2f}, S2={kg_s2}, S3={kg_s3}")
        count += 1
        if count >= 5:
            break

if count == 0:
    print("No se encontraron filas con Kg_S1 > 0")
    # Verifiquemos cual columna tiene Kg reales
    print("\n=== Buscando columna con Kg no nulos ===")
    for col in [86, 87, 88, 89, 90, 91, 13]:
        for fila in range(5, min(200, ws.max_row) + 1):
            val = ws.cell(fila, col).value
            if val and val != 0 and isinstance(val, (int, float)):
                header = ws.cell(4, col).value
                print(f"  Col {col} ({header!r}): Fila {fila} = {val}")
                break

# También verifiquemos la hoja Pesos con detalle
print("\n\n=== HOJA PESOS - primeras 10 filas de datos reales ===")
ws_pesos = wb["Pesos"]
for fila in range(2, 15):
    row = [ws_pesos.cell(fila, c).value for c in range(1, 10)]
    if any(v is not None for v in row):
        print(f"  Fila {fila}: {row}")
