"""Lee todas las hojas del Excel de proyecciones y vuelca su contenido para análisis."""
import openpyxl
import pandas as pd

RUTA = r"D:\Proyecto2026\ACP_DWH\ACP Proyecciones\39. Proyecciones Semana 5 + 5 semanas (Conteo S3).xlsx"

wb = openpyxl.load_workbook(RUTA, data_only=True)
print(f"Hojas: {wb.sheetnames}\n")

for nombre in wb.sheetnames:
    ws = wb[nombre]
    print(f"\n{'='*70}")
    print(f"HOJA: {nombre} ({ws.max_row} filas x {ws.max_column} cols)")
    print(f"{'='*70}")
    # Leer primeras 60 filas para ver estructura
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(60, ws.max_row), values_only=True):
        rows.append(row)
    df = pd.DataFrame(rows)
    # Mostrar solo columnas que tienen datos
    df = df.dropna(how='all', axis=1).dropna(how='all', axis=0)
    print(df.to_string(index=True, header=True))
    print()
