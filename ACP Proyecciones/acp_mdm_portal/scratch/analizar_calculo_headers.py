"""Análisis profundo de la hoja Calculo del Excel - fórmulas y estructura real."""
import openpyxl
import pandas as pd

RUTA = r"D:\Proyecto2026\ACP_DWH\ACP Proyecciones\39. Proyecciones Semana 5 + 5 semanas (Conteo S3).xlsx"

wb = openpyxl.load_workbook(RUTA, data_only=True)
ws = wb["Calculo"]

print(f"Calculo: {ws.max_row} filas x {ws.max_column} cols\n")

# Leer todas las columnas de la primera fila (headers)
print("=== FILA 1 (Headers fila 1) ===")
headers_r1 = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
for i, h in enumerate(headers_r1, 1):
    if h is not None:
        print(f"  Col {i}: {h!r}")

print("\n=== FILA 2 (Headers fila 2) ===")
headers_r2 = [ws.cell(2, c).value for c in range(1, ws.max_column+1)]
for i, h in enumerate(headers_r2, 1):
    if h is not None:
        print(f"  Col {i}: {h!r}")

print("\n=== FILA 3 (Headers fila 3) ===")
headers_r3 = [ws.cell(3, c).value for c in range(1, ws.max_column+1)]
for i, h in enumerate(headers_r3, 1):
    if h is not None:
        print(f"  Col {i}: {h!r}")

# Leer las primeras 10 filas de datos completas
print("\n=== PRIMERAS 10 FILAS DE DATOS (todas las columnas no nulas) ===")
all_rows = []
for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
    all_rows.append(row)

df = pd.DataFrame(all_rows)
# Filtrar columnas que tienen al menos un valor
mask = df.notna().any()
df_filtered = df.loc[:, mask]
print(df_filtered.to_string())

# Buscar columna que tenga "peso" o "baya"
print("\n=== BÚSQUEDA DE COLUMNA 'PESO' en cabeceras ===")
for r in range(1, 5):
    for c in range(1, ws.max_column+1):
        val = ws.cell(r, c).value
        if val and any(k in str(val).lower() for k in ['peso', 'baya', 'kg', 'semana', 'productiva', 'planta', 'madura', 'cosech']):
            print(f"  Fila {r}, Col {c}: {val!r}")

# Leer la segunda hoja del calculo con datos reales
print("\n=== PRIMERA FILA DE DATO REAL (fila ~5 en adelante) ===")
# Buscar primera fila con dato de módulo numérico
for r in range(4, 20):
    row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
    non_none = [v for v in row_vals if v is not None]
    if len(non_none) > 5:
        print(f"\n  Fila {r}:")
        for c_idx, val in enumerate(row_vals, 1):
            if val is not None:
                print(f"    Col {c_idx}: {val!r}")
        break
