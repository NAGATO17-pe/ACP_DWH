"""
paginas/proyecciones.py — Módulo de Proyecciones Fenológicas Six-Week
=======================================================================
UI interactiva para editar la matriz de distribución y visualizar
el volumen proyectado de cosecha semana a semana (W1-W6).
"""

from __future__ import annotations

from datetime import date, datetime, timedelta

import pandas as pd
import streamlit as st

from utils.componentes import estado_vacio_html, banner_aviso
from utils.formato import header_pagina
from utils.motor_proyecciones import (
    MATRIZ_INPUTS_DEFAULT,
    ejecutar_proyeccion,
    extraer_proyeccion_anterior,
    obtener_fechas_disponibles,
    obtener_combinaciones_disponibles,
    verificar_integridad_datos,
    cerrar_matriz,
    validar_matriz_inputs,
    guardar_matriz_inputs,
    cargar_matriz_inputs,
    guardar_proyeccion,
    MARGEN_PESIMISTA,
    MARGEN_OPTIMISTA,
)

# ── Constantes ────────────────────────────────────────────────────────────────
_ESTADOS_ORDENADOS = ["cosechable", "maduras", "cremas", "fase_2", "fase_1", "verdes", "pequena"]
_COLS_SEMANAS = ["W1", "W2", "W3", "W4", "W5", "W6"]



# ── Helpers ───────────────────────────────────────────────────────────────────

def _construir_df_matriz_inputs(fuente: dict | None = None) -> pd.DataFrame:
    """
    Construye el DataFrame que alimenta el data_editor.
    Si `fuente` es None usa MATRIZ_INPUTS_DEFAULT.
    """
    base = fuente if fuente is not None else MATRIZ_INPUTS_DEFAULT
    filas = []
    for est_key in _ESTADOS_ORDENADOS:
        config = base.get(est_key, MATRIZ_INPUTS_DEFAULT.get(est_key, {}))
        fila = {"Estado": est_key.capitalize()}
        for i, col in enumerate(_COLS_SEMANAS):
            val = config.get(i + 1)
            fila[col] = val if val is not None else float("nan")
        filas.append(fila)
    return pd.DataFrame(filas)


def _parsear_matriz_inputs(df_editado: pd.DataFrame) -> dict[str, dict[int, float | None]]:
    matriz: dict[str, dict[int, float | None]] = {}
    for i, est_key in enumerate(_ESTADOS_ORDENADOS):
        if i < len(df_editado):
            row = df_editado.iloc[i]
            matriz[est_key] = {
                j + 1: (None if pd.isna(row[col]) else float(row[col]))
                for j, col in enumerate(_COLS_SEMANAS)
            }
        else:
            matriz[est_key] = MATRIZ_INPUTS_DEFAULT.get(est_key, {})
    return matriz


def _exportar_a_excel(df: pd.DataFrame, nombre_hoja: str = "Desglose") -> bytes:
    """
    Serializa un DataFrame a XLSX en memoria con formato profesional:
    - Encabezado en negrita con fondo slate
    - Auto-ancho de columnas según contenido
    - Cifras numéricas con miles y 0 decimales para columnas Kg
    Retorna los bytes listos para st.download_button.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja[:31]  # Excel limita el nombre de hoja a 31 chars

    cols = list(df.columns)
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="F8FAFC")
        cell.fill = PatternFill("solid", start_color="334155")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    cols_kg = {c for c in cols if str(c).lower().startswith("kg")}
    for _, row in df.iterrows():
        ws.append([row[c] for c in cols])

    for idx, col in enumerate(cols, start=1):
        letra = ws.cell(row=1, column=idx).column_letter
        if col in cols_kg:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx).number_format = "#,##0"
        # Auto-ancho aproximado
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).head(200)])
        ws.column_dimensions[letra].width = min(max(max_len + 2, 10), 28)

    wb.save(buf)
    return buf.getvalue()


def _firma_matriz(matriz: dict[str, dict[int, float | None]]) -> tuple:
    """Hash determinista de la matriz para cachear el reporte de validación."""
    return tuple(
        (est, tuple((w, matriz[est].get(w)) for w in range(1, 7)))
        for est in _ESTADOS_ORDENADOS
        if est in matriz
    )


def _calcular_diff_matriz(
    bd: dict[str, dict[int, float | None]],
    editada: dict[str, dict[int, float | None]],
) -> list[dict]:
    """
    Calcula las diferencias celda a celda entre la matriz cargada desde BD
    y la editada por el usuario. Devuelve solo las celdas con cambio.

    Trata None y NaN como ausente. Tolerancia 1e-6 para floats.
    """
    cambios: list[dict] = []
    for est in _ESTADOS_ORDENADOS:
        bd_est = bd.get(est, {}) or {}
        ed_est = editada.get(est, {}) or {}
        for w in range(1, 7):
            v_bd = bd_est.get(w)
            v_ed = ed_est.get(w)
            v_bd_n = float("nan") if v_bd is None or pd.isna(v_bd) else float(v_bd)
            v_ed_n = float("nan") if v_ed is None or pd.isna(v_ed) else float(v_ed)
            ambos_nan = pd.isna(v_bd_n) and pd.isna(v_ed_n)
            if ambos_nan:
                continue
            if pd.isna(v_bd_n) or pd.isna(v_ed_n) or abs(v_bd_n - v_ed_n) > 1e-6:
                cambios.append({
                    "Estado":  est.capitalize(),
                    "Semana":  f"W{w}",
                    "BD":      None if pd.isna(v_bd_n) else round(v_bd_n, 3),
                    "Editado": None if pd.isna(v_ed_n) else round(v_ed_n, 3),
                    "Δ":       (None if pd.isna(v_bd_n) or pd.isna(v_ed_n)
                                else round(v_ed_n - v_bd_n, 3)),
                })
    return cambios


@st.cache_data(show_spinner=False, max_entries=8)
def _validar_matriz_cacheado(matriz_repr: tuple) -> dict[str, dict]:
    """
    Wrapper cacheado de validar_matriz_inputs.
    `matriz_repr` es la representación hashable de la matriz (misma estructura
    que devuelve `_firma_matriz`). Actúa a la vez como clave de caché y como
    fuente para reconstruir el dict que acepta `validar_matriz_inputs`.
    """
    matriz = {est: {w: v for w, v in semanas} for est, semanas in matriz_repr}
    return validar_matriz_inputs(matriz)


# ── Sección: Cabecera ─────────────────────────────────────────────────────────

def _render_cabecera():
    header_pagina(
        "🌿",
        "Proyecciones Fenológicas · 6 Semanas",
        "Simulación interactiva de cosecha basada en los estados fenológicos actuales",
    )
    
    # --- Widget: Pulso del DWH (Reacción al Event Bus) ---
    try:
        import sys, os
        backend_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "backend"))
        if backend_path not in sys.path:
            sys.path.append(backend_path)
        from servicios.servicio_auditoria import obtener_historial
        historial = obtener_historial(limite=1)
        if historial:
            ultimo = historial[0]
            st.sidebar.markdown("---")
            st.sidebar.markdown("### 📡 Pulso del DWH")
            st.sidebar.info(
                f"**Última actualización:**\n\n"
                f"📦 `{ultimo.get('Tabla_Destino', 'N/A')}`\n\n"
                f"✅ **Estado:** {ultimo.get('Estado', 'N/A')}\n\n"
                f"⏱️ **Hora:** {ultimo.get('Fecha_Inicio', 'N/A')}"
            )
    except Exception as e:
        pass


# ── Sección: Control de fecha y estado ───────────────────────────────────────

def _opciones_validas(df_universo: pd.DataFrame, sel: dict) -> dict:
    """
    Cascada bidireccional: dado el universo de combinaciones de la semana
    y las selecciones actuales (None = 'Todos'), retorna las opciones válidas
    de cada eje aplicando todos los filtros EXCEPTO el del propio eje.

    Esto permite que cualquier filtro pueda quedar en 'Todos' y los demás
    se restringen entre sí por disponibilidad real.
    """
    def filtrar_excepto(campo: str) -> list[str]:
        df = df_universo
        for k, v in sel.items():
            if k != campo and v is not None:
                df = df[df[k] == v]
        return sorted(df[campo].dropna().unique().tolist())

    return {c: filtrar_excepto(c) for c in ["Fundo", "Modulo", "Variedad", "Condicion"]}


# ── Persistencia de filtros en URL (st.query_params) ─────────────────────────
#
# Mapa clave URL ↔ session_state. Centralizado para que añadir un eje sea
# tocar UN solo sitio (cumple "código único"). Las claves URL son cortas
# para que el link sea legible.
_URL_KEYS: dict[str, str] = {
    "w":    "proy_id_tiempo_select",   # semana / id_tiempo (int)
    "fnd":  "proy_sel_fundo",          # nombre del fundo (str | None)
    "mod":  "proy_sel_modulo",         # str del módulo (str | None)
    "var":  "proy_sel_variedad",       # nombre variedad (str | None)
    "cond": "proy_sel_condicion",      # condición compuesta (str | None)
}


def _hidratar_filtros_desde_url() -> None:
    """
    Lee st.query_params UNA vez por sesión y siembra session_state.
    Solo se aplica si la session aún no tiene esas claves seteadas, para no
    pisar selecciones del usuario en reruns posteriores.
    """
    if st.session_state.get("_proy_url_hidratada"):
        return
    qp = st.query_params
    for url_key, ss_key in _URL_KEYS.items():
        if url_key not in qp or ss_key in st.session_state:
            continue
        raw = qp[url_key]
        if url_key == "w":
            try:
                st.session_state[ss_key] = int(raw)
            except (TypeError, ValueError):
                pass
        else:
            st.session_state[ss_key] = raw if raw not in ("", "None") else None
    st.session_state["_proy_url_hidratada"] = True


def _sincronizar_url_filtros(id_tiempo: int, fundo, mod, var, cond) -> None:
    """
    Refleja los filtros activos en la URL. Solo escribe las claves que cambian
    para no disparar reruns innecesarios.
    """
    deseado = {
        "w":    str(id_tiempo) if id_tiempo else "",
        "fnd":  fundo or "",
        "mod":  str(mod) if mod is not None else "",
        "var":  var or "",
        "cond": cond or "",
    }
    qp = st.query_params
    for k, v in deseado.items():
        actual = qp.get(k, "")
        if v and v != actual:
            qp[k] = v
        elif not v and k in qp:
            del qp[k]


def _render_control_fecha() -> tuple[date, int, int | None, str | None, str | None, str | None, bool]:
    # Hidratación una sola vez por sesión: si la URL trae filtros, los siembra
    # en session_state ANTES de instanciar los selectbox para que tomen el valor.
    _hidratar_filtros_desde_url()

    fechas_ids = obtener_fechas_disponibles()

    if not fechas_ids:
        st.error("No se encontraron datos en la base de datos.")
        return date.today(), 0, None, None, None, None, False

    # Convertir IDs a objetos date y mostrar una opción por semana ISO
    opciones_fecha: list[tuple[int, str]] = []
    for fid in fechas_ids:
        try:
            d = datetime.strptime(str(fid), "%Y%m%d").date()
            año, sem_iso, dia_sem = d.isocalendar()
            
            # Calcular lunes y domingo de esa semana para el label
            lunes = d - timedelta(days=dia_sem - 1)
            domingo = lunes + timedelta(days=6)
            
            label = f"Semana {sem_iso} · {año} ({lunes.strftime('%d/%m')} al {domingo.strftime('%d/%m')})"
            opciones_fecha.append((fid, label))
        except Exception:
            continue

    if not opciones_fecha:
        st.error("No se pudieron procesar las fechas de evaluación. Verifica la dimensión de tiempo.")
        return date.today(), 0, None, None, None, None, False

    # ── Fila 1: Semana ────────────────────────────────────────────────
    col_sem, _ = st.columns([2, 6])
    with col_sem:
        seleccion = st.selectbox(
            "📅 Semana de evaluación",
            options=[x[0] for x in opciones_fecha],
            format_func=lambda x: next(o[1] for o in opciones_fecha if o[0] == x),
            key="proy_id_tiempo_select",
        )

    fecha_corte = datetime.strptime(str(seleccion), "%Y%m%d").date()
    id_tiempo = seleccion

    # ── Cargar universo de la semana (cascada bidireccional) ─────────
    df_universo = obtener_combinaciones_disponibles(id_tiempo)

    if df_universo.empty:
        banner_aviso(
            f"La semana **{id_tiempo}** no tiene datos de conteo fenológico. "
            "Sin conteo no es posible ejecutar la proyección. Selecciona otra semana o carga los datos de conteo."
        )
        return fecha_corte, id_tiempo, None, None, None, None, False

    # Selecciones previas (persistentes en session_state, una por eje)
    # Si el usuario cambia de semana, se resetean para evitar selecciones inválidas
    if st.session_state.get("proy_universo_id_tiempo") != id_tiempo:
        st.session_state["proy_universo_id_tiempo"] = id_tiempo
        for k in ("proy_sel_fundo", "proy_sel_modulo", "proy_sel_variedad", "proy_sel_condicion"):
            st.session_state[k] = None

    sel_actual = {
        "Fundo":     st.session_state.get("proy_sel_fundo"),
        "Modulo":    st.session_state.get("proy_sel_modulo"),
        "Variedad":  st.session_state.get("proy_sel_variedad"),
        "Condicion": st.session_state.get("proy_sel_condicion"),
    }

    opts = _opciones_validas(df_universo, sel_actual)

    # Si la selección actual ya no está en las opciones válidas (porque otro filtro
    # la excluyó), la reseteamos a None y refrescamos las opciones.
    cambios = False
    for eje, key in [("Fundo", "proy_sel_fundo"), ("Modulo", "proy_sel_modulo"),
                     ("Variedad", "proy_sel_variedad"), ("Condicion", "proy_sel_condicion")]:
        if sel_actual[eje] is not None and sel_actual[eje] not in opts[eje]:
            st.session_state[key] = None
            sel_actual[eje] = None
            cambios = True
    if cambios:
        opts = _opciones_validas(df_universo, sel_actual)

    # ── Fila 2: 4 dropdowns de filtros ────────────────────────────────
    col_fundo, col_mod, col_var, col_cond = st.columns([2, 1.5, 2, 1.8])

    def _index_o_cero(opciones: list, valor):
        try:
            return ([None] + opciones).index(valor)
        except ValueError:
            return 0

    with col_fundo:
        fundo_sel = st.selectbox(
            "🏞️ Fundo",
            options=[None] + opts["Fundo"],
            format_func=lambda x: x if x is not None else "— Todos —",
            index=_index_o_cero(opts["Fundo"], sel_actual["Fundo"]),
            key="proy_sel_fundo",
        )

    with col_mod:
        mod_str = st.selectbox(
            "🗂 Módulo",
            options=[None] + opts["Modulo"],
            format_func=lambda x: f"Módulo {x}" if x is not None else "— Todos —",
            index=_index_o_cero(opts["Modulo"], sel_actual["Modulo"]),
            key="proy_sel_modulo",
        )
        # Backward compat: convertir a int para el motor (acepta str también)
        try:
            mod_sel = int(mod_str) if mod_str is not None else None
        except (TypeError, ValueError):
            mod_sel = mod_str

    with col_var:
        var_sel = st.selectbox(
            "🌱 Variedad",
            options=[None] + opts["Variedad"],
            format_func=lambda x: x if x is not None else "— Todas —",
            index=_index_o_cero(opts["Variedad"], sel_actual["Variedad"]),
            key="proy_sel_variedad",
        )

    with col_cond:
        cond_sel = st.selectbox(
            "🌿 Condición",
            options=[None] + opts["Condicion"],
            format_func=lambda x: x if x is not None else "— Todas —",
            index=_index_o_cero(opts["Condicion"], sel_actual["Condicion"]),
            key="proy_sel_condicion",
        )

    # ── Verificación de integridad ────────────────────────────────────
    status = verificar_integridad_datos(id_tiempo, mod_sel, var_sel, cond_sel)

    cols_status = st.columns(3)
    with cols_status[0]:
        icon = "✅" if status["conteo"] else "❌"
        st.caption(f"{icon} Conteos Fenol.")
    with cols_status[1]:
        icon = "✅" if status["peladas"] else "⚠️"
        st.caption(f"{icon} Censo de Plantas")
    with cols_status[2]:
        icon = "✅" if status["pesos"] else "⚠️"
        st.caption(f"{icon} Pesos Hist.")

    disponible = status["conteo"]  # El único bloqueante es el conteo

    # ── Línea de contexto con los 4 ejes (siempre visibles) ───────────
    st.caption(
        "**Proyección filtrada por:** "
        f"🏞️ Fundo: **{fundo_sel or 'Todos'}** · "
        f"📦 Módulo: **{mod_sel if mod_sel is not None else 'Todos'}** · "
        f"🌱 Variedad: **{var_sel or 'Todas'}** · "
        f"🌿 Condición: **{cond_sel or 'Todas'}**"
    )

    # Reflejar filtros activos en la URL (link compartible / refresh-safe).
    _sincronizar_url_filtros(id_tiempo, fundo_sel, mod_sel, var_sel, cond_sel)

    return fecha_corte, id_tiempo, mod_sel, var_sel, cond_sel, fundo_sel, disponible


def _render_panel_configuracion() -> tuple[dict, bool]:
    """
    Renderiza el editor de la matriz de maduración + botones cargar/guardar.

    Los márgenes optimista/pesimista NO son configurables desde la UI: usan
    los defaults `MARGEN_PESIMISTA` y `MARGEN_OPTIMISTA` del Excel.

    Returns:
        (matriz_inputs, matriz_valida) — la matriz parseada y un flag que
        indica si pasó las validaciones (True = se puede ejecutar).
    """
    with st.expander("⚙️ Configuración de Escenarios", expanded=True):
        # Cabecera con botones de carga/persistencia
        col_titulo, col_load, col_save = st.columns([5, 1.2, 1.2])
        with col_titulo:
            st.markdown(
                "### 1. Entradas Manuales (Matriz de Maduración)\n"
                "Edita los coeficientes conocidos. Deja las celdas vacías para que se auto-calculen con `1 - Σ`."
            )
        with col_load:
            cargar_clk = st.button(
                "📥 Cargar BD",
                use_container_width=True,
                help="Carga la última matriz guardada en Config.Parametros_Pipeline",
                key="proy_cargar_matriz_btn",
            )
        with col_save:
            # El guardado real se hace tras parsear el editor (más abajo)
            guardar_placeholder = st.empty()

        # Si el usuario pidió cargar, sustituimos la fuente por la matriz BD
        fuente_inicial: dict | None = None
        if cargar_clk:
            data_bd = cargar_matriz_inputs()
            if data_bd:
                fuente_inicial = data_bd
                st.session_state["proy_matriz_bd_cargada"] = data_bd
                st.toast("Matriz cargada desde BD.", icon="📥")
            else:
                banner_aviso("No hay matriz guardada en BD todavía.")
        elif "proy_matriz_bd_cargada" in st.session_state:
            fuente_inicial = st.session_state["proy_matriz_bd_cargada"]

        df_inputs = _construir_df_matriz_inputs(fuente_inicial)

        df_editado = st.data_editor(
            df_inputs,
            hide_index=True,
            use_container_width=True,
            key="proy_matriz_editor",
            column_config={
                "Estado": st.column_config.TextColumn("Estado Fenológico", disabled=True, width="medium"),
                **{
                    col: st.column_config.NumberColumn(col, min_value=0.0, max_value=1.0, step=0.01, format="%.3f")
                    for col in _COLS_SEMANAS
                },
            },
        )

        matriz_inputs = _parsear_matriz_inputs(df_editado)

        # ── Diff visual: matriz BD vs editada (#16) ─────────────────────
        # Solo se muestra si el usuario cargó una matriz desde BD y la modificó.
        matriz_bd_ref = st.session_state.get("proy_matriz_bd_cargada")
        if matriz_bd_ref:
            cambios_diff = _calcular_diff_matriz(matriz_bd_ref, matriz_inputs)
            if cambios_diff:
                with st.expander(
                    f"✏️ {len(cambios_diff)} celda(s) modificadas vs BD",
                    expanded=False,
                ):
                    df_diff = pd.DataFrame(cambios_diff)
                    st.dataframe(
                        df_diff,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Estado":   st.column_config.TextColumn("Estado", width="medium"),
                            "Semana":   st.column_config.TextColumn("Semana", width="small"),
                            "BD":       st.column_config.NumberColumn("Valor BD", format="%.3f"),
                            "Editado":  st.column_config.NumberColumn("Editado", format="%.3f"),
                            "Δ":        st.column_config.NumberColumn("Delta", format="%+.3f"),
                        },
                    )

        # ── Validación de la matriz (cacheada por firma) ────────────────
        firma = _firma_matriz(matriz_inputs)
        reporte = _validar_matriz_cacheado(firma)

        errores_globales = [
            f"**{est.capitalize()}** — {err}"
            for est, info in reporte.items() for err in info["errores"]
        ]
        avisos_globales = [
            f"**{est.capitalize()}** — {adv}"
            for est, info in reporte.items() for adv in info["advertencias"]
        ]
        # #9 — set de estados con error para resaltarlos visualmente arriba
        estados_con_error = {est for est, info in reporte.items() if info["errores"]}

        matriz_valida = not errores_globales

        if errores_globales:
            st.error(
                "❌ La matriz tiene errores que impiden ejecutar la proyección:\n\n- "
                + "\n- ".join(errores_globales)
            )
            if estados_con_error:
                st.caption(
                    "🔴 Filas a corregir: "
                    + ", ".join(f"**{e.capitalize()}**" for e in sorted(estados_con_error))
                )
                # Mini-tabla diagnóstico: muestra solo las filas con error con
                # iconos por celda para que el usuario localice el problema rápido.
                df_diag = df_inputs[df_inputs["Estado"].str.lower().isin(estados_con_error)].copy()
                for est in estados_con_error:
                    sem = matriz_inputs.get(est, {})
                    for w in range(1, 7):
                        v = sem.get(w)
                        marca = ""
                        if v is not None and (v < 0 or v > 1):
                            marca = "🔴"  # fuera de rango
                        df_diag.loc[df_diag["Estado"].str.lower() == est, f"W{w}"] = (
                            f"{marca} {v:.3f}" if v is not None and marca else
                            (f"{v:.3f}" if v is not None else "—")
                        )
                # Si la suma del estado excede 1, marcamos toda la fila con ⚠️
                df_diag.insert(1, "Σ inputs", [
                    f"⚠️ {reporte[est.lower()]['suma']:.3f}"
                    if reporte[est.lower()]["suma"] > 1.0 + 1e-6
                    else f"{reporte[est.lower()]['suma']:.3f}"
                    for est in df_diag["Estado"]
                ])
                st.dataframe(df_diag, hide_index=True, use_container_width=True)
        if avisos_globales:
            with st.expander(f"ℹ️ {len(avisos_globales)} advertencia(s) informativas", expanded=False):
                for adv in avisos_globales:
                    st.caption(f"• {adv}")

        # Botón de guardar (se renderiza ahora que ya tenemos matriz_inputs)
        with guardar_placeholder.container():
            if st.button(
                "💾 Guardar BD",
                use_container_width=True,
                disabled=not matriz_valida,
                help="Persiste la matriz actual en Config.Parametros_Pipeline",
                key="proy_guardar_matriz_btn",
            ):
                ok = guardar_matriz_inputs(matriz_inputs)
                if ok:
                    st.toast("Matriz guardada en BD.", icon="✅")

        # ── Matriz cerrada (Solo lectura) ───────────────────────────────
        st.markdown("### 2. Matriz Calculada (Cierre 100%)")
        matriz_cerrada = cerrar_matriz(matriz_inputs)

        df_cerrada = pd.DataFrame([
            {"Estado": est.capitalize(), **{f"W{i+1}": matriz_cerrada[est][i] for i in range(6)}}
            for est in _ESTADOS_ORDENADOS
            if est in matriz_cerrada
        ])

        st.dataframe(
            df_cerrada,
            hide_index=True,
            use_container_width=True,
            column_config={
                "Estado": st.column_config.TextColumn("Estado Fenológico", width="medium"),
                **{f"W{i+1}": st.column_config.ProgressColumn(f"W{i+1}", min_value=0, max_value=1, format="%.3f") for i in range(6)}
            }
        )

    return matriz_inputs, matriz_valida



# ── Sección: KPIs ─────────────────────────────────────────────────────────────

def _render_kpis(kpis: dict, margen_pes: float = MARGEN_PESIMISTA, margen_opt: float = MARGEN_OPTIMISTA):
    from utils.formato import crear_tarjeta_kpi

    total_base = kpis.get("total_base", 0)
    total_opt = kpis.get("total_opt", 0)
    total_pes = kpis.get("total_pes", 0)
    total_plantas = kpis.get("total_plantas", 0)
    kg_planta = kpis.get("kg_por_planta", 0)
    u_cubiertas = kpis.get("unidades_cubiertas", 0)
    u_totales = kpis.get("unidades_totales", 1)
    eficiencia = (u_cubiertas / u_totales) * 100 if u_totales else 0

    st.markdown('<div class="kpi-container">', unsafe_allow_html=True)

    tips = {
        "tn":   "Suma de Kg base proyectados W1-W6 / 1000. Refleja los filtros activos.",
        "kg_p": "Kg base totales / Plantas totales. Productividad media por planta para la ventana de 6 semanas.",
        "ef":   f"Unidades (Mód×Turno×Válv×Var) con datos de Peladas / Total = {u_cubiertas} / {u_totales}. "
                "Las que no tienen censo usan el fallback de 1500 plantas.",
        "pl":   "Suma de plantas estimadas para todas las unidades incluidas en la proyección.",
        "opt":  f"Kg base × {margen_opt:.4f} (factor MARGEN_OPTIMISTA del Excel).",
        "pes":  f"Kg base × {margen_pes:.4f} (factor MARGEN_PESIMISTA del Excel).",
    }

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(crear_tarjeta_kpi("Toneladas Proyectadas", f"{total_base/1000:,.1f} Tn", "🫐", "info", tips["tn"]), unsafe_allow_html=True)
    with col2:
        st.markdown(crear_tarjeta_kpi("Kilos por Planta", f"{kg_planta:.2f} kg", "🪴", "success", tips["kg_p"]), unsafe_allow_html=True)
    with col3:
        st.markdown(crear_tarjeta_kpi("Eficiencia Censo", f"{eficiencia:.1f}%", "📊", "info", tips["ef"]), unsafe_allow_html=True)
    with col4:
        st.markdown(crear_tarjeta_kpi("Plantas Totales", f"{total_plantas:,.0f}", "🚜", "info", tips["pl"]), unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    # Segunda fila para escenarios
    col_a, col_b, _ = st.columns([1, 1, 2])
    delta_opt = f"+{((total_opt / total_base) - 1) * 100:.1f}%" if total_base else "—"
    delta_pes = f"-{(1 - (total_pes / total_base)) * 100:.1f}%" if total_base else "—"

    with col_a:
        st.markdown(
            f"<div title='{tips['opt']}' style='padding:10px 0;'>"
            f"<div style='font-size:0.65rem;color:#8fa897;font-weight:700;text-transform:uppercase;"
            f"letter-spacing:0.8px;margin-bottom:4px;'>Escenario Optimista</div>"
            f"<div style='font-size:1.2rem;font-weight:700;color:#2db87a;"
            f"font-family:JetBrains Mono,monospace;'>{total_opt:,.0f} kg</div>"
            f"<div style='font-size:0.75rem;color:#2db87a;opacity:0.75;'>{delta_opt}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )
    with col_b:
        st.markdown(
            f"<div title='{tips['pes']}' style='padding:10px 0;'>"
            f"<div style='font-size:0.65rem;color:#8fa897;font-weight:700;text-transform:uppercase;"
            f"letter-spacing:0.8px;margin-bottom:4px;'>Escenario Pesimista</div>"
            f"<div style='font-size:1.2rem;font-weight:700;color:#e8a020;"
            f"font-family:JetBrains Mono,monospace;'>{total_pes:,.0f} kg</div>"
            f"<div style='font-size:0.75rem;color:#e8a020;opacity:0.75;'>{delta_pes}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )



# ── Sección: Gráfico de curva de cosecha ─────────────────────────────────────

def _aplicar_ventana_semanas(
    df_semanal: pd.DataFrame,
    df_anterior: pd.DataFrame,
    rango: tuple[int, int],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Filtra ambos DataFrames al sub-rango (semana_ini, semana_fin).
    Helper único reutilizado por el gráfico y la comparación con histórico.
    """
    s_ini, s_fin = rango
    df_s = df_semanal[df_semanal["semana"].between(s_ini, s_fin)].copy()
    if df_anterior.empty:
        return df_s, df_anterior
    df_a = df_anterior[df_anterior["semana"].between(s_ini, s_fin)].copy()
    return df_s, df_a


def _render_grafico_cosecha(df_semanal: pd.DataFrame, df_anterior: pd.DataFrame):
    """
    Gráfico de cosecha proyectada con:
    - Banda de incertidumbre sombreada (área entre pesimista y optimista)
    - Barras base con etiquetas de valor en Tn
    - Anotación automática del pico semanal
    - Eje Y en Toneladas (escala más legible para volúmenes operativos)
    """
    try:
        import plotly.graph_objects as go
    except ImportError:
        banner_aviso("Instala plotly para ver el gráfico: `pip install plotly`")
        return

    if df_semanal.empty:
        banner_aviso("No hay datos en la ventana seleccionada.")
        return

    # Trabajamos en Toneladas para el eje Y (más legible que kg).
    df = df_semanal.copy()
    df["base_t"] = df["kg_base"] / 1000.0
    df["opt_t"]  = df["kg_optimista"] / 1000.0
    df["pes_t"]  = df["kg_pesimista"] / 1000.0
    etiquetas = df["base_t"].apply(lambda v: f"{v:.1f} T")

    fig = go.Figure()

    # ── Banda de incertidumbre (pesimista → optimista) ───────────────
    # Línea inferior invisible que sirve de ancla para el fill.
    fig.add_trace(go.Scatter(
        x=df["semana_label"], y=df["pes_t"],
        line=dict(width=0), hoverinfo="skip",
        showlegend=False,
    ))
    fig.add_trace(go.Scatter(
        x=df["semana_label"], y=df["opt_t"],
        name="Banda Pesimista–Optimista",
        line=dict(width=0),
        fill="tonexty",
        fillcolor="rgba(232, 160, 32, 0.10)",
        hovertemplate="<b>%{x}</b><br>Optimista: %{y:.2f} T<extra></extra>",
    ))

    # ── Barras base con etiquetas de valor ────────────────────────────
    fig.add_trace(go.Bar(
        x=df["semana_label"], y=df["base_t"],
        name="Base",
        marker_color="rgba(232, 160, 32, 0.75)",
        marker_line_color="#e8a020",
        marker_line_width=1,
        text=etiquetas, textposition="outside",
        textfont=dict(color="#e8f0ec", size=11, family="Inter, sans-serif"),
        hovertemplate="<b>%{x}</b><br>Base: %{y:.2f} T<extra></extra>",
    ))

    # ── Líneas optimista/pesimista (referencia explícita) ─────────────
    fig.add_trace(go.Scatter(
        x=df["semana_label"], y=df["opt_t"],
        name="Optimista",
        mode="lines+markers",
        line=dict(color="#2db87a", width=2, dash="dot"),
        marker=dict(size=6),
        hovertemplate="<b>%{x}</b><br>Optimista: %{y:.2f} T<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=df["semana_label"], y=df["pes_t"],
        name="Pesimista",
        mode="lines+markers",
        line=dict(color="#EF4444", width=2, dash="dot"),
        marker=dict(size=6),
        hovertemplate="<b>%{x}</b><br>Pesimista: %{y:.2f} T<extra></extra>",
    ))

    # ── Proyección anterior (si existe) ───────────────────────────────
    if not df_anterior.empty and df_anterior["kg_anterior"].sum() > 0:
        fig.add_trace(go.Scatter(
            x=df_anterior["semana_label"],
            y=df_anterior["kg_anterior"] / 1000.0,
            name="Proyección Anterior (BD)",
            mode="lines+markers",
            line=dict(color="#8fa897", width=1.5, dash="dash"),
            marker=dict(size=5, symbol="diamond"),
            opacity=0.75,
            hovertemplate="<b>%{x}</b><br>Anterior: %{y:.2f} T<extra></extra>",
        ))

    # ── Anotación de pico semanal ─────────────────────────────────────
    if df["base_t"].sum() > 0:
        idx_pico = df["base_t"].idxmax()
        x_pico = df.loc[idx_pico, "semana_label"]
        y_pico = float(df.loc[idx_pico, "base_t"])
        # Etiqueta corta (ej. "W3" del label "W3 (15/05)")
        sem_corto = str(x_pico).split(" ", 1)[0]
        fig.add_annotation(
            x=x_pico, y=y_pico,
            text=f"📈 Pico {sem_corto} · {y_pico:.1f} T",
            showarrow=True, arrowhead=2, arrowsize=1, arrowwidth=1,
            arrowcolor="#e8a020",
            ax=0, ay=-40,
            bgcolor="rgba(232, 160, 32, 0.15)",
            bordercolor="#e8a020", borderwidth=1, borderpad=4,
            font=dict(color="#e8f0ec", size=11, family="Inter, sans-serif"),
        )

    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(5, 15, 8, 0.0)",
        plot_bgcolor="rgba(5, 15, 8, 0.2)",
        title=dict(
            text="Curva de Cosecha Proyectada — 6 Semanas",
            font=dict(family="Inter, sans-serif", size=18, color="#e8f0ec"),
            x=0, y=0.95,
        ),
        xaxis=dict(
            title="Semana de Cosecha",
            color="#8fa897",
            gridcolor="rgba(255,255,255,0.03)",
            showline=False,
        ),
        yaxis=dict(
            title="Toneladas (T)",
            color="#8fa897",
            gridcolor="rgba(255,255,255,0.03)",
            showline=False,
            zeroline=False,
            tickformat=".1f",
        ),
        legend=dict(
            orientation="h",
            yanchor="bottom", y=1.02,
            xanchor="right", x=1,
            bgcolor="rgba(0,0,0,0)",
            font=dict(color="#8fa897", size=11),
        ),
        hovermode="x unified",
        margin=dict(l=0, r=0, t=80, b=0),
        height=470,
    )


    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


# ── Sección: Comparación con proyección anterior ──────────────────────────────

def _render_comparacion(df_semanal: pd.DataFrame, df_anterior: pd.DataFrame, id_tiempo: int, res_completo: dict):
    if df_anterior.empty or df_anterior["kg_anterior"].sum() == 0:
        st.markdown(f"""
            <div style="background:rgba(26,46,30,0.4); border:1px solid rgba(255,255,255,0.07); border-radius:14px; padding:32px; text-align:center; margin-bottom:12px;">
                <div style="font-size:2.5rem; margin-bottom:12px;">📊</div>
                <h4 style="margin:0; color:#e8f0ec; opacity:0.85;">Sin registros previos</h4>
                <p style="color:#8fa897; font-size:0.9rem; margin-top:8px;">No existe una proyección guardada para esta semana.</p>
            </div>
        """, unsafe_allow_html=True)
        
        # Botón estándar fuera de HTML para máxima confiabilidad
        if st.button("💾 Guardar Proyección Oficial", type="primary", use_container_width=True):
            with st.spinner("Persistiendo datos en el histórico..."):
                res_save = guardar_proyeccion(res_completo["df_detalle"], id_tiempo)
                if not res_save.get("errores"):
                    st.cache_data.clear() # Limpiar caché para que lea los nuevos datos guardados
                    st.success(f"✅ ¡Guardado con éxito! ({res_save.get('insertados', 0)} filas)")
                    st.balloons()
                    
                    # --- Notificación Desacoplada al Central Command ---
                    try:
                        import sys, os
                        backend_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "backend"))
                        if backend_path not in sys.path:
                            sys.path.append(backend_path)
                        from servicios.event_bus import bus
                        bus.projection_saved.send('PORTAL_PROYECCIONES', filas=res_save.get('insertados', 0))
                        st.toast("📡 Señal enviada: [PROJECTION_SAVED]", icon="📡")
                    except Exception as e:
                        st.toast("⚠️ No se pudo notificar al Central Command", icon="⚠️")

                    st.rerun()

                else:
                    st.error(f"❌ Error: {res_save['errores'][0]}")
        return




    df_comp = df_semanal[["semana", "semana_label", "kg_base"]].merge(
        df_anterior[["semana", "kg_anterior"]], on="semana", how="outer"
    ).fillna(0)

    df_comp["diferencia_kg"] = df_comp["kg_base"] - df_comp["kg_anterior"]
    df_comp["variacion_pct"] = (
        ((df_comp["kg_base"] - df_comp["kg_anterior"]) / df_comp["kg_anterior"].replace(0, float("nan"))) * 100
    ).round(1)

    df_comp = df_comp.rename(columns={
        "semana_label": "Semana",
        "kg_base": "Kg Actual",
        "kg_anterior": "Kg Anterior (BD)",
        "diferencia_kg": "Diferencia (kg)",
        "variacion_pct": "Variación (%)",
    })

    # ── NUEVO: Gráfico de Barras Comparativo ──────────────────────────────────
    import plotly.graph_objects as go
    fig_comp = go.Figure()
    
    # Barras de Kg Anterior (Histórico)
    fig_comp.add_trace(go.Bar(
        x=df_comp["Semana"],
        y=df_comp["Kg Anterior (BD)"],
        name="Anterior",
        marker_color="rgba(148, 163, 184, 0.4)", # Slate
        marker_line=dict(color="rgba(148, 163, 184, 0.8)", width=1)
    ))
    
    # Barras de Kg Actual
    fig_comp.add_trace(go.Bar(
        x=df_comp["Semana"],
        y=df_comp["Kg Actual"],
        name="Actual",
        marker_color="#2db87a",
        marker_line=dict(color="#4dd99a", width=1)
    ))
    
    fig_comp.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(5, 15, 8, 0.2)",
        barmode="group",
        height=350,
        margin=dict(l=0, r=0, t=10, b=10),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        yaxis=dict(title="Kilogramos", gridcolor="rgba(255,255,255,0.05)"),
        xaxis=dict(title=None)
    )
    st.plotly_chart(fig_comp, use_container_width=True, config={"displayModeBar": False})

    total_actual = df_comp["Kg Actual"].sum()
    total_anterior = df_comp["Kg Anterior (BD)"].sum()
    delta_total = total_actual - total_anterior
    delta_pct = ((delta_total / total_anterior) * 100) if total_anterior else 0


    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(
            f"<div style='padding:8px 0;'>"
            f"<div style='font-size:0.65rem;color:#8fa897;font-weight:700;text-transform:uppercase;"
            f"letter-spacing:0.8px;margin-bottom:4px;'>Total Actual</div>"
            f"<div style='font-size:1.2rem;font-weight:700;color:#2db87a;"
            f"font-family:JetBrains Mono,monospace;'>{total_actual:,.0f} kg</div>"
            f"</div>",
            unsafe_allow_html=True,
        )
    with col2:
        st.markdown(
            f"<div style='padding:8px 0;'>"
            f"<div style='font-size:0.65rem;color:#8fa897;font-weight:700;text-transform:uppercase;"
            f"letter-spacing:0.8px;margin-bottom:4px;'>Total Anterior</div>"
            f"<div style='font-size:1.2rem;font-weight:700;color:#8fa897;"
            f"font-family:JetBrains Mono,monospace;'>{total_anterior:,.0f} kg</div>"
            f"</div>",
            unsafe_allow_html=True,
        )
    with col3:
        color_var = "#2db87a" if delta_total >= 0 else "#EF4444"
        st.markdown(
            f"<div style='padding:8px 0;'>"
            f"<div style='font-size:0.65rem;color:#8fa897;font-weight:700;text-transform:uppercase;"
            f"letter-spacing:0.8px;margin-bottom:4px;'>Variación Total</div>"
            f"<div style='font-size:1.2rem;font-weight:700;color:{color_var};"
            f"font-family:JetBrains Mono,monospace;'>{delta_total:+,.0f} kg</div>"
            f"<div style='font-size:0.75rem;color:{color_var};opacity:0.75;'>{delta_pct:+.1f}%</div>"
            f"</div>",
            unsafe_allow_html=True,
        )

    st.dataframe(
        df_comp[["Semana", "Kg Actual", "Kg Anterior (BD)", "Diferencia (kg)", "Variación (%)"]],
        use_container_width=True,
        hide_index=True,
    )
    
    # Sobrescritura con confirmación explícita (2 pasos)
    col_chk, col_btn = st.columns([3, 2])
    with col_chk:
        confirmar_overwrite = st.checkbox(
            "Confirmo sobrescribir la proyección oficial guardada",
            key="proy_chk_overwrite",
            help="Al sobrescribir, los kg actualmente guardados en Fact_Proyecciones "
                 "para esta semana y filtros se reemplazan por los de pantalla.",
        )
    with col_btn:
        if st.button(
            "🔄 Sobrescribir Proyección",
            key="btn_overwrite",
            disabled=not confirmar_overwrite,
            use_container_width=True,
        ):
            with st.spinner("Sobrescribiendo histórico…"):
                res_save = guardar_proyeccion(res_completo["df_detalle"], id_tiempo)
            if not res_save.get("errores"):
                st.cache_data.clear()
                st.success(f"✅ Datos actualizados ({res_save.get('insertados', 0)} filas).")
                
                # --- Notificación Desacoplada al Central Command ---
                try:
                    import sys, os
                    backend_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "backend"))
                    if backend_path not in sys.path:
                        sys.path.append(backend_path)
                    from servicios.event_bus import bus
                    bus.projection_saved.send('PORTAL_PROYECCIONES', filas=res_save.get('insertados', 0), overwrite=True)
                    st.toast("📡 Señal enviada: [PROJECTION_SAVED] (Update)", icon="📡")
                except Exception as e:
                    st.toast("⚠️ No se pudo notificar al Central Command", icon="⚠️")
                
                st.rerun()
            else:
                st.error(f"❌ Error: {res_save['errores'][0]}")




# ── Sección: KPIs por Fundo y comparación Orgánico vs Convencional ───────────

def _render_kpis_por_fundo(
    df_detalle: pd.DataFrame,
    fundo_sel: str | None = None,
    cond_sel: str | None = None,
):
    """
    Muestra una tabla compacta con kg totales por Fundo (W1-W6 + total) y un KPI
    de Δ Orgánico vs Convencional.

    Reglas de visibilidad:
    - Si df_detalle no tiene columna 'fundo' o está vacío → no se muestra nada.
    - Si fundo_sel está fijo (un solo fundo) → la tabla por Fundo es ruido (1 fila),
      se omite. El bloque Δ Org vs Conv se mantiene salvo que también haya
      condición fija.
    """
    if df_detalle.empty or "fundo" not in df_detalle.columns:
        return

    mostrar_tabla_fundo = fundo_sel is None
    # Clasificación explícita Org/Conv usa la columna `certificacion`
    # (no la derivamos del string `condicion`).
    mostrar_delta_cond = cond_sel is None and "certificacion" in df_detalle.columns

    if not mostrar_tabla_fundo and not mostrar_delta_cond:
        return  # Nada útil que mostrar en esta sección

    st.markdown("#### 🏞️ Desempeño por Fundo")

    if mostrar_tabla_fundo:
        # Pivot Fundo × semana ordenada por número (no por label) para que
        # la sparkline respete W1→W6.
        df_fundo = (
            df_detalle.groupby(["fundo", "semana"])["kg_base"]
            .sum()
            .reset_index()
            .sort_values(["fundo", "semana"])
        )
        pivot = (
            df_fundo.pivot(index="fundo", columns="semana", values="kg_base")
            .fillna(0)
            .reindex(columns=range(1, 7), fill_value=0)
        )
        # Una columna lista (sparkline) y columnas Wn por separado para vista numérica.
        pivot_view = pd.DataFrame(index=pivot.index)
        pivot_view["Tendencia"] = pivot.values.tolist()
        for w in range(1, 7):
            pivot_view[f"W{w}"] = pivot[w].round(0).astype(int)
        pivot_view["Total Tn"] = (pivot.sum(axis=1) / 1000.0).round(1)
        pivot_view = pivot_view.sort_values("Total Tn", ascending=False)

        st.dataframe(
            pivot_view.reset_index().rename(columns={"fundo": "Fundo"}),
            use_container_width=True,
            hide_index=True,
            column_config={
                "Fundo":     st.column_config.TextColumn("Fundo", width="medium"),
                "Tendencia": st.column_config.LineChartColumn(
                    "Tendencia W1-W6",
                    help="Curva de cosecha proyectada por semana",
                    width="small",
                ),
                **{f"W{w}": st.column_config.NumberColumn(f"W{w}", format="%d") for w in range(1, 7)},
                "Total Tn":  st.column_config.NumberColumn(
                    "Total Tn",
                    help="Suma de kg base / 1000",
                    format="%.1f",
                ),
            },
        )

    # Δ Orgánico vs Convencional — clasificación explícita por Certificacion
    # (Dim_Condicion_Cultivo.Certificacion). Robusto a variaciones de nombre.
    if mostrar_delta_cond:
        cert_norm = df_detalle["certificacion"].astype(str).str.strip().str.lower()
        es_organico = cert_norm.eq("organico") | cert_norm.eq("orgánico")

        organico = df_detalle.loc[es_organico, "kg_base"].sum()
        convencional = df_detalle.loc[~es_organico & cert_norm.ne("sin certificación"), "kg_base"].sum()
        total = organico + convencional

        if total > 0:
            col_a, col_b, col_c = st.columns(3)
            delta_pct = ((organico - convencional) / convencional * 100) if convencional > 0 else 0
            with col_a:
                st.markdown(
                    f"<div style='padding:8px 0;'>"
                    f"<div style='font-size:0.65rem;color:#8fa897;font-weight:700;text-transform:uppercase;"
                    f"letter-spacing:0.8px;margin-bottom:4px;'>🌱 Orgánico</div>"
                    f"<div style='font-size:1.2rem;font-weight:700;color:#2db87a;"
                    f"font-family:JetBrains Mono,monospace;'>{organico/1000:,.1f} Tn</div>"
                    f"<div style='font-size:0.75rem;color:#2db87a;opacity:0.75;'>{organico/total*100:.1f}%</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            with col_b:
                st.markdown(
                    f"<div style='padding:8px 0;'>"
                    f"<div style='font-size:0.65rem;color:#8fa897;font-weight:700;text-transform:uppercase;"
                    f"letter-spacing:0.8px;margin-bottom:4px;'>🧪 Convencional</div>"
                    f"<div style='font-size:1.2rem;font-weight:700;color:#e8a020;"
                    f"font-family:JetBrains Mono,monospace;'>{convencional/1000:,.1f} Tn</div>"
                    f"<div style='font-size:0.75rem;color:#e8a020;opacity:0.75;'>{convencional/total*100:.1f}%</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            with col_c:
                color_d = "#2db87a" if delta_pct >= 0 else "#EF4444"
                st.markdown(
                    f"<div style='padding:8px 0;'>"
                    f"<div style='font-size:0.65rem;color:#8fa897;font-weight:700;text-transform:uppercase;"
                    f"letter-spacing:0.8px;margin-bottom:4px;'>Δ Org vs Conv</div>"
                    f"<div style='font-size:1.2rem;font-weight:700;color:{color_d};"
                    f"font-family:JetBrains Mono,monospace;'>{(organico-convencional)/1000:+.1f} Tn</div>"
                    f"<div style='font-size:0.75rem;color:{color_d};opacity:0.75;'>{delta_pct:+.1f}%</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )


def _render_comparativa_organico_convencional(df_detalle: pd.DataFrame, cond_sel: str | None):
    """
    Gráfico de barras agrupadas Orgánico vs Convencional por semana W1-W6.
    Solo se muestra cuando cond_sel es None ('Todas').

    La clasificación se hace por la columna `certificacion`
    (Dim_Condicion_Cultivo.Certificacion), no por sub-string del nombre,
    para resistir variaciones tipográficas ("Orgánico" con tilde, "ORG", etc.).
    """
    if cond_sel is not None or df_detalle.empty or "certificacion" not in df_detalle.columns:
        return

    try:
        import plotly.graph_objects as go
    except ImportError:
        return

    df = df_detalle.copy()
    cert_norm = df["certificacion"].astype(str).str.strip().str.lower()

    def _clasificar(c: str) -> str:
        if c in ("organico", "orgánico"):
            return "Orgánico"
        if c == "sin certificación":
            return "Sin condición"
        return "Convencional"

    df["grupo"] = cert_norm.map(_clasificar)

    df_agg = df.groupby(["semana_label", "grupo"])["kg_base"].sum().reset_index()
    if df_agg.empty or df_agg["kg_base"].sum() == 0:
        return

    st.markdown("#### 🌱 Comparativa Orgánico vs Convencional")

    fig = go.Figure()
    colores = {"Orgánico": "#2db87a", "Convencional": "#e8a020", "Sin condición": "#8fa897"}
    for grupo in ["Orgánico", "Convencional", "Sin condición"]:
        sub = df_agg[df_agg["grupo"] == grupo]
        if sub.empty:
            continue
        fig.add_trace(go.Bar(
            x=sub["semana_label"],
            y=sub["kg_base"],
            name=grupo,
            marker_color=colores[grupo],
        ))

    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(5, 15, 8, 0.2)",
        barmode="group",
        height=350,
        margin=dict(l=0, r=0, t=10, b=10),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        yaxis=dict(title="Kilogramos", gridcolor="rgba(255,255,255,0.05)"),
        xaxis=dict(title=None),
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


# ── Sección: Vistas analíticas (Heatmap + Treemap) ──────────────────────────

def _render_heatmap_modulo_semana(df_detalle: pd.DataFrame):
    """
    Heatmap Módulo × Semana en kg base.
    Ordena módulos por total descendente para que los más productivos queden
    arriba. El color usa una escala secuencial Plasma (oscura → brillante).
    """
    if df_detalle.empty or "modulo" not in df_detalle.columns:
        return
    try:
        import plotly.graph_objects as go
    except ImportError:
        return

    pivot = (
        df_detalle.groupby(["modulo", "semana"])["kg_base"]
        .sum()
        .reset_index()
        .pivot(index="modulo", columns="semana", values="kg_base")
        .fillna(0)
        .sort_index()
    )
    if pivot.empty:
        return

    # Reordenar por total descendente: mayor productividad arriba
    pivot["__total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("__total", ascending=True).drop(columns="__total")
    semanas = sorted(pivot.columns.tolist())
    pivot = pivot[semanas]
    valores_t = pivot.values / 1000.0  # mostramos Toneladas

    fig = go.Figure(data=go.Heatmap(
        z=valores_t,
        x=[f"W{w}" for w in semanas],
        y=[f"Mod {m}" for m in pivot.index],
        colorscale="Plasma",
        colorbar=dict(
            title=dict(text="Tn", font=dict(color="#8fa897", size=11)),
            tickfont=dict(color="#8fa897", size=10),
        ),
        hovertemplate="<b>%{y}</b> · <b>%{x}</b><br>%{z:.2f} Tn<extra></extra>",
        zmin=0,
    ))

    altura = max(280, min(28 * len(pivot.index) + 80, 700))
    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(5, 15, 8, 0.2)",
        title=dict(
            text=f"Heatmap Módulo × Semana ({len(pivot.index)} módulos)",
            font=dict(family="Inter, sans-serif", size=15, color="#e8f0ec"),
            x=0, y=0.96,
        ),
        xaxis=dict(side="top", color="#8fa897", gridcolor="rgba(255,255,255,0.03)"),
        yaxis=dict(color="#8fa897", gridcolor="rgba(255,255,255,0.03)"),
        margin=dict(l=0, r=0, t=60, b=20),
        height=altura,
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


def _render_treemap_distribucion(df_detalle: pd.DataFrame):
    """
    Treemap jerárquico: Fundo → Condición → Módulo → Variedad por kg_base.
    Muestra de un vistazo dónde se concentra la cosecha proyectada y
    permite drill-down haciendo clic en cualquier rectángulo.
    """
    if df_detalle.empty:
        return
    try:
        import plotly.express as px
    except ImportError:
        return

    # Solo incluimos columnas presentes para no romper si falta alguna
    candidatos = ["fundo", "condicion", "modulo", "variedad"]
    path = [c for c in candidatos if c in df_detalle.columns]
    if not path:
        return

    df = df_detalle.copy()
    # Treemap ignora valores 0/negativos: filtramos para evitar warnings
    df = df[df["kg_base"] > 0]
    if df.empty:
        return

    # Pre-renombramos columnas para que el treemap muestre labels capitalizados
    rename = {
        "fundo": "Fundo", "condicion": "Condición",
        "modulo": "Módulo", "variedad": "Variedad",
    }
    df = df.rename(columns=rename)
    df["Módulo"] = df["Módulo"].apply(lambda m: f"Mod {m}")
    path_labels = [rename[c] for c in path]

    df_agg = df.groupby(path_labels, as_index=False)["kg_base"].sum()

    fig = px.treemap(
        df_agg,
        path=[px.Constant("Total")] + path_labels,
        values="kg_base",
        color="kg_base",
        color_continuous_scale=[
            (0.0, "#0d2014"), (0.4, "#e8a020"), (1.0, "#2db87a"),
        ],
        hover_data={"kg_base": ":.0f"},
    )
    fig.update_traces(
        texttemplate="<b>%{label}</b><br>%{value:,.0f} kg<br>%{percentParent:.1%}",
        textfont=dict(family="Inter, sans-serif", size=12, color="#e8f0ec"),
        hovertemplate="<b>%{label}</b><br>%{value:,.0f} kg (%{percentRoot:.1%} del total)<extra></extra>",
        marker=dict(line=dict(color="rgba(15,23,42,0.7)", width=1)),
    )
    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        title=dict(
            text="Distribución jerárquica de la cosecha proyectada",
            font=dict(family="Inter, sans-serif", size=15, color="#e8f0ec"),
            x=0, y=0.97,
        ),
        margin=dict(l=0, r=0, t=50, b=10),
        height=480,
        coloraxis_colorbar=dict(
            title=dict(text="kg", font=dict(color="#8fa897", size=11)),
            tickfont=dict(color="#8fa897", size=10),
        ),
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


def _render_vistas_analiticas(df_detalle: pd.DataFrame):
    """
    Sección colapsable que agrupa visualizaciones densas: Heatmap (vista
    cuantitativa) y Treemap (vista de composición). Pestañas para alternar
    entre ambas sin saturar la pantalla.
    """
    if df_detalle.empty:
        return
    with st.expander("🗺️ Vistas analíticas (Heatmap · Treemap)", expanded=False):
        tab_heat, tab_tree = st.tabs(["🔥 Heatmap Módulo × Semana", "🌳 Treemap jerárquico"])
        with tab_heat:
            st.caption(
                "Densidad de cosecha por módulo y semana. Filas ordenadas de menor a mayor "
                "total. Colores brillantes = mayor volumen."
            )
            _render_heatmap_modulo_semana(df_detalle)
        with tab_tree:
            st.caption(
                "Composición del total proyectado por jerarquía Fundo → Condición → "
                "Módulo → Variedad. Haz clic en un rectángulo para hacer drill-down."
            )
            _render_treemap_distribucion(df_detalle)


# ── Sección: Desglose granular ────────────────────────────────────────────────

def _render_desglose(df_detalle: pd.DataFrame):
    with st.expander("🔍 Desglose Granular de la Proyección", expanded=False):
        if df_detalle.empty:
            estado_vacio_html(icono="📊", titulo="Sin desglose disponible")
            return

        # Detectar si tenemos las nuevas columnas de trazabilidad
        tiene_fundo = "fundo" in df_detalle.columns
        tiene_condicion = "condicion" in df_detalle.columns

        # Opciones de agrupación dinámicas según trazabilidad disponible
        opciones_grupo = []
        if tiene_fundo:
            opciones_grupo.append("Fundo")
        opciones_grupo.append("Módulo")
        opciones_grupo.append("Módulo × Variedad")
        opciones_grupo.append("Módulo × Turno × Variedad")
        if tiene_fundo:
            opciones_grupo.append("Trazabilidad completa (Fundo→Cond→Mod→Var)")
        opciones_grupo.append("Detalle (Mod × Turno × Válvula)")

        col_group, _ = st.columns([4, 4])
        with col_group:
            agrupar_por = st.radio(
                "Nivel de Agrupación",
                options=opciones_grupo,
                index=opciones_grupo.index("Módulo × Variedad"),
                horizontal=False,
                key="proy_agrupar_radio",
            )

        df_proc = df_detalle.copy()
        agg_cols = {
            "kg_base": ("kg_base", "sum"),
            "kg_pesimista": ("kg_pesimista", "sum"),
            "kg_optimista": ("kg_optimista", "sum"),
        }

        if agrupar_por == "Fundo":
            df_proc = (
                df_proc.groupby(["fundo", "semana", "semana_label", "fecha_semana"])
                .agg(**agg_cols)
                .reset_index()
                .sort_values(["fundo", "semana"])
            )
            cols_show = ["fundo", "semana_label", "kg_base", "kg_pesimista", "kg_optimista"]
        elif agrupar_por == "Módulo":
            df_proc = (
                df_proc.groupby(["modulo", "semana", "semana_label", "fecha_semana"])
                .agg(**agg_cols)
                .reset_index()
                .sort_values(["modulo", "semana"])
            )
            cols_show = ["modulo", "semana_label", "kg_base", "kg_pesimista", "kg_optimista"]
        elif agrupar_por == "Módulo × Variedad":
            df_proc = (
                df_proc.groupby(["modulo", "variedad", "semana", "semana_label", "fecha_semana"])
                .agg(**agg_cols)
                .reset_index()
                .sort_values(["modulo", "variedad", "semana"])
            )
            cols_show = ["modulo", "variedad", "semana_label", "kg_base", "kg_pesimista", "kg_optimista"]
        elif agrupar_por == "Módulo × Turno × Variedad":
            df_proc = (
                df_proc.groupby(["modulo", "turno", "variedad", "semana", "semana_label", "fecha_semana"])
                .agg(**agg_cols)
                .reset_index()
                .sort_values(["modulo", "turno", "variedad", "semana"])
            )
            cols_show = ["modulo", "turno", "variedad", "semana_label", "kg_base", "kg_pesimista", "kg_optimista"]
        elif agrupar_por.startswith("Trazabilidad"):
            df_proc = (
                df_proc.groupby(["fundo", "condicion", "modulo", "variedad",
                                  "semana", "semana_label", "fecha_semana"])
                .agg(**agg_cols)
                .reset_index()
                .sort_values(["fundo", "condicion", "modulo", "variedad", "semana"])
            )
            cols_show = ["fundo", "condicion", "modulo", "variedad", "semana_label",
                         "kg_base", "kg_pesimista", "kg_optimista"]
        else:  # Detalle completo
            cols_base = (["fundo"] if tiene_fundo else []) + (["condicion"] if tiene_condicion else [])
            df_proc = df_proc.sort_values(cols_base + ["modulo", "turno", "valvula", "variedad", "semana"])
            cols_show = cols_base + ["modulo", "turno", "valvula", "variedad",
                                      "semana_label", "kg_base", "kg_pesimista", "kg_optimista"]

        df_mostrar = df_proc[cols_show].rename(columns={
            "fundo": "Fundo",
            "condicion": "Condición",
            "modulo": "Módulo",
            "turno": "Turno",
            "valvula": "Válvula",
            "variedad": "Variedad",
            "semana_label": "Semana",
            "kg_base": "Kg Base",
            "kg_pesimista": "Kg Pesimista",
            "kg_optimista": "Kg Optimista",
        })

        st.dataframe(
            df_mostrar,
            use_container_width=True,
            hide_index=True,
        )

        # Export a Excel — el archivo lleva el nivel de agrupación en el nombre
        try:
            xlsx_bytes = _exportar_a_excel(df_mostrar, nombre_hoja="Proyeccion")
            slug = (
                agrupar_por.lower()
                .replace(" × ", "_")
                .replace(" ", "_")
                .replace("(", "")
                .replace(")", "")
                .replace("→", "_")
            )
            st.download_button(
                "📥 Exportar a Excel",
                data=xlsx_bytes,
                file_name=f"proyeccion_6w_{slug}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=False,
                help="Descarga el desglose actual en formato Excel con encabezados formateados.",
            )
        except ImportError:
            st.caption("ℹ️ Instala `openpyxl` para habilitar la exportación a Excel.")



# ── Punto de entrada ──────────────────────────────────────────────────────────

def render():
    _render_cabecera()

    fecha_corte, id_tiempo, mod_sel, var_sel, cond_sel, fundo_sel, disponible = _render_control_fecha()


    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    matriz, matriz_valida = _render_panel_configuracion()
    # Márgenes operativos: defaults fijos del Excel (no editables en UI).
    margen_pes, margen_opt = MARGEN_PESIMISTA, MARGEN_OPTIMISTA

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    col_btn, _ = st.columns([2, 6])
    with col_btn:
        puede_ejecutar = disponible and matriz_valida
        ayuda = (
            "Selecciona una fecha con datos disponibles para ejecutar"
            if not disponible
            else "Corrige los errores de la matriz para habilitar"
            if not matriz_valida
            else "Calcula la proyección con los inputs actuales"
        )
        ejecutar = st.button(
            "⚡ Ejecutar Proyección",
            type="primary",
            disabled=not puede_ejecutar,
            use_container_width=True,
            help=ayuda,
        )

    if ejecutar:
        with st.spinner("Calculando proyecciones fenológicas…"):
            resultado = ejecutar_proyeccion(
                id_tiempo,
                matriz,
                margen_pes,
                margen_opt,
                modulo=mod_sel,
                variedad=var_sel,
                condicion=cond_sel,
                fundo=fundo_sel,
            )

        if resultado["df_semanal"].empty:
            st.error("❌ No se encontraron datos para generar la proyección con estos filtros. Verifica que existan registros de conteo fenológico para la selección.")
        else:
            st.session_state["proy_resultado"] = resultado
            st.session_state["proy_id_tiempo"] = id_tiempo
            # Guardamos la firma del filtro con el que se calculó (incluye fundo)
            st.session_state["proy_filtro"] = (id_tiempo, mod_sel, var_sel, cond_sel, fundo_sel)
            st.toast("🚀 Proyección generada con éxito", icon="✅")


    if "proy_resultado" in st.session_state and not st.session_state["proy_resultado"]["df_semanal"].empty:
        res = st.session_state["proy_resultado"]
        id_t = st.session_state.get("proy_id_tiempo", id_tiempo)

        # Advertir si el usuario cambió los filtros desde el último cálculo
        filtro_calculado = st.session_state.get("proy_filtro")
        filtro_actual = (id_tiempo, mod_sel, var_sel, cond_sel, fundo_sel)

        if filtro_calculado and filtro_calculado != filtro_actual:
            banner_aviso(
                "Los filtros han cambiado desde el último cálculo. "
                "Pulsa **Ejecutar Proyección** para actualizar los resultados."
            )

        st.markdown("---")
        st.markdown("### 📊 Resultados de la Proyección")

        _render_kpis(res["kpis"], margen_pes=margen_pes, margen_opt=margen_opt)

        # Extraer proyección anterior con TODOS los filtros activos (incluido Fundo)
        # para que la comparación refleje exactamente el mismo subconjunto que
        # se proyectó. Si el usuario dejó algo en 'Todos', el motor recibe None
        # y no aplica ese filtro.
        df_anterior = extraer_proyeccion_anterior(
            id_t,
            modulo=mod_sel,
            variedad=var_sel,
            condicion=cond_sel,
            fundo=fundo_sel,
        )




        # ── Ventana visible (W1-W6) ───────────────────────────────────
        # Slider que filtra TODOS los visualizadores siguientes a un sub-rango.
        # No re-ejecuta el motor: solo recorta el resultado para análisis what-if.
        # Para extender más allá de W6 se requiere validar W7-W11 operacionalmente.
        col_slider, col_info = st.columns([3, 4])
        with col_slider:
            rango_sem = st.slider(
                "Ventana visible (semanas)",
                min_value=1, max_value=6, value=(1, 6), step=1,
                key="proy_rango_semanas",
                help="Recorta los gráficos y KPIs al sub-rango. No recalcula el motor.",
            )
        with col_info:
            st.caption(
                f"Mostrando **W{rango_sem[0]}–W{rango_sem[1]}** "
                f"({rango_sem[1] - rango_sem[0] + 1} semanas de 6). "
                "El cálculo siempre se hace sobre las 6 semanas completas; "
                "el slider solo controla la vista."
            )

        # Aplicamos la ventana una sola vez y reutilizamos en todos los renders
        df_semanal_v, df_anterior_v = _aplicar_ventana_semanas(
            res["df_semanal"], df_anterior, rango_sem
        )
        df_detalle_v = res["df_detalle"][
            res["df_detalle"]["semana"].between(rango_sem[0], rango_sem[1])
        ]

        _render_grafico_cosecha(df_semanal_v, df_anterior_v)

        # KPIs por Fundo + comparativa Orgánico vs Convencional.
        # _render_kpis_por_fundo se auto-oculta si Fundo está fijo y Cond también.
        st.markdown("---")
        _render_kpis_por_fundo(df_detalle_v, fundo_sel=fundo_sel, cond_sel=cond_sel)
        _render_comparativa_organico_convencional(df_detalle_v, cond_sel)

        # Vistas analíticas densas (colapsadas por defecto para no saturar)
        _render_vistas_analiticas(df_detalle_v)

        st.markdown("---")
        st.markdown("### 🔄 Comparación con Proyección Anterior")
        _render_comparacion(df_semanal_v, df_anterior_v, id_t, res)


        _render_desglose(df_detalle_v)
